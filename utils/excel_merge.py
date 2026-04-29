from __future__ import annotations

import io
from dataclasses import dataclass
from typing import List, Optional, Tuple


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except ImportError as e:
        raise RuntimeError("未安装 openpyxl，请在项目目录执行: pip install openpyxl") from e


@dataclass(frozen=True)
class MergeXlsxOptions:
    """
    将多个工作簿按顺序合并到单个 Sheet（默认：合并各文件的 active sheet）。
    """

    skip_duplicate_header: bool = True
    align_by_header: bool = True
    header_case_insensitive: bool = False
    column_merge_mode: str = "union"  # union: 补齐并追加额外列；intersection: 仅保留共同列


def merge_xlsx_bytes(
    files: List[Tuple[str, bytes]],
    *,
    opts: Optional[MergeXlsxOptions] = None,
) -> bytes:
    """
    合并 Excel（.xlsx/.xlsm）为一个 .xlsx：按传入顺序将每个文件的 active sheet 逐行追加。

    默认按表头对齐（align_by_header=True）：
    - 先确定统一表头（以第一个文件为基准）
    - 后续文件缺失列自动补空
    - 后续文件新增列自动追加到末尾
    """
    openpyxl = _require_openpyxl()
    from copy import copy

    if not files:
        raise RuntimeError("未提供要合并的表格文件")
    opts = opts or MergeXlsxOptions()

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "合并"

    def _is_non_empty(v) -> bool:
        return v is not None and str(v).strip() != ""

    def _norm_header(v) -> str:
        s = "" if v is None else str(v).strip()
        return s

    def _canon_header(v: str) -> str:
        return v.lower() if opts.header_case_insensitive else v

    # 预读取：识别每个 sheet 的表头行/数据起始行，并构建统一列顺序
    prepared = []
    first_order: List[str] = []
    union_order: List[str] = []
    common_headers = None
    canon_to_label: dict[str, str] = {}
    for idx, (_fname, b) in enumerate(files):
        wb = openpyxl.load_workbook(io.BytesIO(b), data_only=False)
        ws = wb.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        header_row = 1
        for r in range(1, max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(_is_non_empty(v) for v in vals):
                header_row = r
                break

        header_vals_raw = [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]
        while header_vals_raw and not _is_non_empty(header_vals_raw[-1]):
            header_vals_raw.pop()
        if not header_vals_raw:
            continue

        headers = []
        canon_headers = []
        for i, v in enumerate(header_vals_raw, start=1):
            h = _norm_header(v) or f"列{i}"
            headers.append(h)
            canon_headers.append(_canon_header(h))

        # 同文件内重名表头去重（避免映射冲突）
        seen = {}
        uniq_headers = []
        uniq_canons = []
        for h, c in zip(headers, canon_headers):
            n = seen.get(c, 0) + 1
            seen[c] = n
            if n == 1:
                uniq_headers.append(h)
                uniq_canons.append(c)
            else:
                uniq_headers.append(f"{h}_{n}")
                uniq_canons.append(f"{c}__{n}")
        headers = uniq_headers
        canon_headers = uniq_canons

        if not union_order:
            union_order = canon_headers[:]
            first_order = canon_headers[:]
            common_headers = set(canon_headers)
        else:
            for c in canon_headers:
                if c not in union_order:
                    union_order.append(c)  # 多出来的列放后面
            if common_headers is not None:
                common_headers &= set(canon_headers)

        for h, c in zip(headers, canon_headers):
            canon_to_label.setdefault(c, h)

        prepared.append(
            {
                "idx": idx,
                "ws": ws,
                "max_row": max_row,
                "header_row": header_row,
                "headers": headers,
                "canon_headers": canon_headers,
            }
        )

        # 复制列宽（仅首个有效文件）
        if idx == 0:
            try:
                for col, dim in ws.column_dimensions.items():
                    if dim.width:
                        out_ws.column_dimensions[col].width = dim.width
            except Exception:
                pass

    if not prepared:
        raise RuntimeError("上传的表格为空或无法读取")

    mode = (opts.column_merge_mode or "union").strip().lower()
    if mode not in ("union", "intersection"):
        mode = "union"
    if mode == "intersection":
        ch = common_headers or set()
        final_canons = [c for c in first_order if c in ch]
    else:
        final_canons = union_order[:]
    if not final_canons:
        raise RuntimeError("表头无可合并列，请检查上传表格")

    # 写统一表头
    for c, canon in enumerate(final_canons, start=1):
        out_ws.cell(row=1, column=c, value=canon_to_label.get(canon, canon))

    wrote_data = False
    for item in prepared:
        ws = item["ws"]
        max_row = item["max_row"]
        header_row = item["header_row"]
        canon_headers = item["canon_headers"]

        # 建立 “输入列索引 -> 输出列索引” 映射
        out_pos = {c: i + 1 for i, c in enumerate(final_canons)}
        col_map = {
            in_idx + 1: out_pos[c]
            for in_idx, c in enumerate(canon_headers)
            if c in out_pos
        }

        start_row = header_row + 1
        if not opts.skip_duplicate_header:
            start_row = header_row

        for r in range(start_row, max_row + 1):
            # 全空行跳过
            row_has_value = False
            for in_c in col_map.keys():
                v = ws.cell(row=r, column=in_c).value
                if _is_non_empty(v):
                    row_has_value = True
                    break
            if not row_has_value:
                continue

            out_r = out_ws.max_row + 1
            wrote_data = True
            for in_c, out_c in col_map.items():
                src = ws.cell(row=r, column=in_c)
                dst = out_ws.cell(row=out_r, column=out_c, value=src.value)
                try:
                    if src.has_style:
                        dst.font = copy(src.font)
                        dst.fill = copy(src.fill)
                        dst.border = copy(src.border)
                        dst.alignment = copy(src.alignment)
                        dst.number_format = src.number_format
                        dst.protection = copy(src.protection)
                        dst._style = copy(src._style)
                except Exception:
                    pass

    if not wrote_data:
        raise RuntimeError("上传的表格为空或无法读取")

    out = io.BytesIO()
    out_wb.save(out)
    return out.getvalue()

